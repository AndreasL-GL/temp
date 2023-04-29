import openpyxl
import os
from datetime import date
import datetime
import json

if __name__ == '__main__':
    from date_functions import get_first_and_last_week_of_month, get_month_from_year_week, get_week_numbers
    from date_functions import get_month_from_week, get_month_name_from_number, first_day_from_week
    import io
    from save_file_on_error import save_file_on_error
    
else:
    from functions.Excel.date_functions import get_first_and_last_week_of_month, get_month_from_year_week, get_week_numbers
    from functions.Excel.date_functions import get_month_from_week, get_month_name_from_number, first_day_from_week
    import io
    from functions.Excel.get_excel_file import download_excel_file
    from tools_get_files import save_file_on_error

def get_dictionary_from_dagbok_sheet(sheet):
    "Creates a dictionary from a dagbook excel sheet"
    dag = [x.value for x in [x[0] for x in sheet['C4:C16']][::2]]
    fastighet = [x.value for x in [x[0] for x in sheet['A4:A16']][::2]]
    trees = [x.value for x in [x[0] for x in sheet['L5:L17']][::2]]
    name = [x for x in [x.value for x in [x[0] for x in sheet['A20:A26']][::2]] if x]
    
    if any(name): name = [name[0] for x in range(0,7)]
    else: name = [sheet.title for x in range(0,7)]
    veckodag = [x.value for x in [x[0] for x in sheet['E20:E26']]]
    start_kl = [x.value for x in [x[0] for x in sheet['F20:F26']]]
    slut_kl = [x.value for x in [x[0] for x in sheet['G20:G26']]]
    rast_tim = [x.value for x in [x[0] for x in sheet['H20:H26']]]
    arbetspost1 = [x.value for x in [x[0] for x in sheet['I20:I26']]]  #I19 sheet['I19'].value
    arbetspost2 = [x.value for x in [x[0] for x in sheet['J20:J26']]]  #J19 sheet['J19'].value
    arbetspost3 = [x.value for x in [x[0] for x in sheet['K20:K26']]]  #K19 sheet['K19'].value
    arbetspost4 = [x.value for x in [x[0] for x in sheet['L20:L26']]]  #L19 sheet['L19'].value
    sa_arbetstid = [x.value for x in [x[0] for x in sheet['M20:M26']]] #M19 sheet['M19'].value
    resa_from = [x.value for x in [x[0] for x in sheet['F29:F35']]] #F28 sheet['F28'].value
    resa_till = [x.value for x in [x[0] for x in sheet['H29:H35']]] #H28 sheet['H28'].value
    km = [x.value for x in [x[0] for x in sheet['K29:K35']]] #K28 sheet['K28'].value
    restid = [x.value for x in [x[0] for x in sheet['L29:L35']]] #L28 sheet['L28'].value

    js = [
        {
            "beskrivning":dag, 
            "fastighet":fastighet, 
            "trees":trees,
            "personalnamn":name_1,
            "Start Kl":[str(start_kl_1) if type(start_kl_1)!=datetime.datetime else start_kl_1.strftime("%H:%M")][0].replace('.',':').replace(',',':'),
            "Slut Kl":[str(slut_kl_1) if type(slut_kl_1)!=datetime.datetime else slut_kl_1.strftime("%H:%M")][0].replace('.',':').replace(',',':'),
            ["Rast" if "rast" in str(sheet["H19"].value).lower() else sheet["H19"].value][0]:rast_1,
            "Veckodag":veckodag_1,
            sheet['I19'].value:arbetspost1_1,
            sheet['J19'].value:arbetspost2_1,
            sheet['K19'].value:arbetspost3_1,
            sheet['L19'].value:arbetspost4_1,
            "Övrigt":0,
            "Resa från":resa_from_1,
            "Resa till":resa_till_1,
            "km":km_1,
            "Restid":restid_1,
            "Oberäknad tid":0
            } 
        for dag, fastighet, trees, name_1, start_kl_1, slut_kl_1, rast_1,veckodag_1, arbetspost1_1,
        arbetspost2_1,arbetspost3_1,arbetspost4_1,resa_from_1,resa_till_1,km_1,restid_1
        in zip(fastighet,dag,trees, name,start_kl,slut_kl,rast_tim,veckodag, arbetspost1,arbetspost2,arbetspost3,
            arbetspost4, resa_from, resa_till, km, restid)]
    
    js2 = []
    i=1
    poster = ['SOS ledare','Lastbil','Avant med förare','Skotning','Skotare','Mark Arb', 'Platschef', 'Träd- besiktning','Trädbesiktare',\
        'Byggmöten','Arborist']
    
    for item in js:
        
        if "Rast" not in item.keys():
            all_values_empty = not item[sheet['I19'].value] \
            and not item[sheet['J19'].value] \
            and not item[sheet['K19'].value] \
            and not item[sheet['L19'].value] \
            and not item[sheet['H19'].value]
        else:
            all_values_empty = not item[sheet['I19'].value] \
        and not item[sheet['J19'].value] \
        and not item[sheet['K19'].value] \
        and not item[sheet['L19'].value]
        if "Rast" not in item.keys(): item["Rast"] = 0
    
        
        if not item[sheet['I19'].value]: item[sheet['I19'].value] = 0
        if not item[sheet['J19'].value]: item[sheet['J19'].value] = 0
        if not item[sheet['K19'].value]: item[sheet['K19'].value] = 0
        if not item[sheet['L19'].value]: item[sheet['L19'].value] = 0
        try:
            if sheet['I19'].value and sheet['I19'].value not in poster: item["Övrigt"]=item["Övrigt"]+int(item[sheet['I19'].value])
            if sheet['J19'].value and sheet['J19'].value not in poster: item["Övrigt"]=item["Övrigt"]+int(item[sheet['J19'].value])
            if sheet['K19'].value and sheet['K19'].value not in poster: item["Övrigt"]=item["Övrigt"]+int(item[sheet['K19'].value])
            if sheet['L19'].value and sheet['L19'].value not in poster: item["Övrigt"]=item["Övrigt"]+int(item[sheet['L19'].value])
        except Exception as e:
            print(e)
        if "time" in str(type(item["Start Kl"]).__repr__): item[sheet['F19'].value] = item[sheet['F19'].value].strftime("%H:%M")
        if "time" in str(type(item["Slut Kl"]).__repr__): item[sheet['G19'].value] = item[sheet['G19'].value].strftime("%H:%M")
        # Om alla värdena är tomma och det ändå står tid skrivet, så sätts variabeln Oberäknad tid till
        # decimalvärdet för tidsskillnaden, dvs. 16:30 - 08:00 = 8.5
        # Med en timma rast: 8.5 - 1 = 7.5
        slut_kl_bool="Slut Kl" in item.keys() and item["Slut Kl"]!="None" and all_values_empty
        start_kl_bool="Start Kl" in item.keys() and item["Start Kl"]!="None" and all_values_empty
        if start_kl_bool and slut_kl_bool:

            if ':' in item['Slut Kl'] and ':' in item['Start Kl']: 
                hours = int(item["Slut Kl"].split(':')[0])-int(item["Start Kl"].split(':')[0])
                minutes = int(item["Slut Kl"].split(':')[1])-int(item["Start Kl"].split(':')[1])
            elif ':' in item['Slut Kl'] and not ':' in item['Start Kl']:
                hours = int(item["Slut Kl"].split(':')[0]) -int(item["Start Kl"])
                minutes = 0
            elif ':' in item['Start Kl'] and not ':' in item['Slut Kl']:
                hours = int(item["Slut Kl"]) -int(item["Start Kl"].split(':')[0])
                minutes = 0
            else: 
                hours = int(item["Slut Kl"]) -int(item["Start Kl"])
                minutes = 0
            minutes = minutes/60
            if item["Rast"]: rast = float(item["Rast"])
            else: rast = 0
            
            item["Oberäknad tid"] = hours+minutes + rast

            
        iso_week = sheet["J3"].value

        if sheet["J2"].value:
            item["Datum"] = date.fromisocalendar(sheet['J2'].value, sheet['J3'].value, i).strftime("%Y-%m-%d")
        elif not iso_week: iso_week = 15
        else: item["Datum"] = date.fromisocalendar(2023, iso_week, i).strftime("%Y-%m-%d")
        i+=1
        js2.append(item)
    poster = ['SOS ledare','Lastbil','Avant med förare','Skotning','Skotare','Mark Arb', 'Platschef', 'Träd- besiktning','Trädbesiktare',\
        'Byggmöten','Arborist']
    poster_i_js=[sheet['I19'].value,sheet['J19'].value,sheet['K19'].value,sheet['L19'].value]

    other = sum([sum([int(js2[i][x]) for x in poster_i_js if x not in poster]) for i in range(len(js2))])
        


    results={
            "info":{
                "Kontrakt nr":sheet['G2'].value,
                "Bandel":sheet['G3'].value,
                "År":[sheet['J2'].value if sheet['J2'].value else datetime.datetime.now().year][0],
                "Vecka":sheet['J3'].value,
                "Lag nr":sheet['L2'].value,
                "Dag":sheet['L3'].value,
                "Dagboksnamn":sheet['D1'].value,
                "Övrig arbetstid":other,
                "Boktitel":sheet.title
            },
            "poster":js2
        }
    results["info"]["Första dag i veckan"] = first_day_from_week(results["info"]["År"], results["info"]["Vecka"])
    print("Första dag i veckan: ", results["info"]["Första dag i veckan"])
    results["info"]["Sammanställning"]=sum([float(item[sheet['I19'].value]) + float(item[sheet['J19'].value]) + float(item[sheet['K19'].value]) + float(item[sheet['L19'].value]) for item in results['poster']])
    results["info"]["Månad"] = get_month_from_week(results["info"]["År"], results["info"]["Vecka"])
    return results
def split_daterange_elements(element):
    letters = ""
    numbers = ""
    for char in element:
        if char.isalpha():
            letters += char
        elif char.isdigit():
            numbers += char
    return (letters, numbers)

@save_file_on_error
def convert_file_to_workbook(bytefile):
    wb = openpyxl.load_workbook(bytefile)

    wb,filename = call_functions(wb)
    file_data = io.BytesIO()
    wb.save(file_data)
    wb.close()
    file_data.seek(0)
    return file_data,filename
    
    
def collect_workbook(items):
    filename = items["info"]["Månad"] + " - Sammanställning - Trädexperterna"+".xlsx"
    l = download_excel_file("TrdexperternaApplikationer",filename)
    if l.status_code == 200:
        file_datas = io.BytesIO(l.content)
        # file_data= io.BytesIO
        # file_data.write(file_datas.getvalue())
        # file_data.seek(0)
        
        wb = openpyxl.load_workbook(file_datas)
    else: 
        wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__),'template2.xlsx'))
    return wb, filename

def call_functions(wb, sheet=None):
    if not sheet:sheet = wb.active
    items = get_dictionary_from_dagbok_sheet(sheet)
    with open(os.path.join(os.path.join(os.path.dirname(__file__),'items trädexperterna'),items["info"]["Boktitel"]+' '+items["info"]["Första dag i veckan"]+'.json'),'w') as f:
            jitems = stringify_dict(items)
            json.dump(jitems,f)
    wb,filename = collect_workbook(items)
    wb = enter_items_into_sheet(wb,items)
    return wb,filename
    
    
def enter_items_into_sheet(wb, items): 
    """Manually enters all cells into a sheet."""
    sheet = wb.active
    # Deklarera variabler
    year = items['info']["År"]
    if not year: year=2023
    week = items['info']["Vecka"]
    unknown="Okänd"
    
    # SET DATE CELLS
    first_and_last_day = get_first_and_last_week_of_month(year,get_month_from_year_week(year,week))
    daterange = excel_range_to_list("F5:AS5")
    dates = get_date_range(*first_and_last_day)
    if len(dates)!=len(daterange): len(dates)
    print(first_and_last_day)
    print(week)
    print(items["info"]["Månad"])
    for i in range(len(dates)):
        sheet[daterange[i]] = int(dates[i].strftime('%d'))
        
    # SET INDEX NUMBER OF ITEM POSTS BASED ON DATE
    new_lst = []
    for element in daterange:
        new_tuple=split_daterange_elements(element)
        new_lst.append(new_tuple)
    daterange=new_lst
    if any([int(x[1]) > 200 for x in daterange]):
        [print(x[1]) > 200 for x in daterange]
    date_index = {date:index[0] for date, index in zip(dates,daterange)}
    
    
    # SET WEEK CELLS
    week_cells = ["F4","M4","T4","AA4","AH4"]
    weeks = get_week_numbers(dates)
    weeks.sort()
    print(weeks)
    print(week_cells)
    week_cells=week_cells[:len(weeks)]
    for i,cell in enumerate(week_cells):
        sheet[cell] = "V " + str(weeks[i])
        
    # SET MONTH AND TITLE
    sheet["B3"] = get_month_name_from_number(get_month_from_year_week(year=year,week=week)) + " " + "2013"
    sheet["E2"] = "Fakturaunderlag " + get_month_name_from_number(get_month_from_year_week(year=year,week=week)) + " " + "2013"
    
    
    # SET BANDEL
    if 'Bandel' in items['info'].keys():
        for cell in excel_range_to_list("B4:D4"):
            if sheet[cell] == items['info']['Bandel']: break
            elif sheet[cell].value: break
            else: 
                sheet[cell] = items['info']['Bandel']
                break
    
    # SET TIMES OF POSTS

    for i in range(len(items['poster'])):
        year1,month,day = items['poster'][i]['Datum'].split('-')
        
        year1,month,day = int(year1),int(month),int(day)
        items['poster'][i]['column_index'] = date_index[datetime.date(year1,month,day)]

        
        # SET RESTID
    if 'Platschef' in items['poster'][0].keys() or 'Trädbesiktare' in items['poster'][0].keys() or 'Träd- besiktning' in items['poster'][0].keys():
        if any([x['Restid'] for x in items['poster'] if x]):
            if sheet["AA149"].value:
                sheet["AA149"] = sheet["AA149"].value+sum([x["Restid"] for x in items['poster'] if x['Restid']])
            else:
                sheet["AA149"] =  sum([x["Restid"] for x in items['poster'] if x['Restid']])
    else:
        if any([x['Restid'] for x in items['poster'] if x['Restid']]):
            if sheet["AA155"].value:
                sheet["AA155"] =sheet["AA155"].value+ sum([x["Restid"] for x in items['poster'] if x['Restid']])
            else:
                sheet["AA155"]  = sum([x["Restid"] for x in items['poster'] if x['Restid']])
        
        
    # SET ARBORIST TIMES AND KM
    if 'Arborist' in items['poster'][0].keys():
        if not sheet['A36'].value:
            if any([x['Arborist'] for x in items['poster'] if x!='0']):
                for cell in excel_range_to_list("A27:A36"):
                    if sheet[cell].value: 
                        if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                    if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                        sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                        index_number = cell[1:]
                    
                        
                        cell_index = [item['column_index'] + index_number for item in items['poster']]
                        for index, arboristtimmar in zip(cell_index,[x['Arborist'] for x in items['poster']]):
                            sheet[index] = arboristtimmar

                        break
        else:
            if any([x['Arborist'] for x in items['poster'] if x!='0']):
                for cell in excel_range_to_list("A38:A47"):
                    if sheet[cell].value: 
                        if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                    if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                        sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                        index_number = cell[1:]
                        cell_index = [item['column_index'] + index_number for item in items['poster']]
                        for index, arboristtimmar in zip(cell_index,[x['Arborist'] for x in items['poster']]):
                            sheet[index] = arboristtimmar
                        break
        if any([x['km'] for x in items['poster'] if x!='0']):
                for cell in excel_range_to_list("A49:A67"):
                    if sheet[cell].value: 
                        if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                    if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                        sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]

                        index_number = cell[1:]
                        cell_index = [item['column_index'] + index_number for item in items['poster']]
                        for index, km in zip(cell_index,[x['km'] for x in items['poster']]):
                            sheet[index] = km
                        break
                
    # SET TRÄDBESIKTARE TIME AND KM
    if 'Trädbesiktare' in items['poster'][0].keys():
        if any([x['Trädbesiktare'] for x in items['poster'] if x!='0']):
            for cell in excel_range_to_list("A10:A16"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['Trädbesiktare'] for x in items['poster']]):
                        sheet[index] = besiktartimmar
                    break
            
        if any([x['km'] for x in items['poster'] if x!='0']):
                for cell in excel_range_to_list("A18:A24"):
                    if sheet[cell].value: 
                        if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                    if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                        sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                        
                        index_number = cell[1:]
                        cell_index = [item['column_index'] + index_number for item in items['poster']]
                        for index, km in zip(cell_index,[x['km'] for x in items['poster']]):
                            sheet[index] = km

                        break
    if 'Träd- besiktning' in items['poster'][0].keys() and any([x['Träd- besiktning'] for x in items['poster'] if x!='0']):
        for cell in excel_range_to_list("A10:A16"):
            if sheet[cell].value: 
                if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
            if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                index_number = cell[1:]
  
                cell_index = [item['column_index'] + index_number for item in items['poster']]
                for index, besiktartimmar in zip(cell_index,[x['Träd- besiktning'] for x in items['poster']]):
                    sheet[index] = besiktartimmar
                break
            
        if any([x['km'] for x in items['poster'] if x!='0']):
                for cell in excel_range_to_list("A18:A24"):
                    if sheet[cell].value: 
                        if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                    if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                        sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                        
                        index_number = cell[1:]
                        cell_index = [item['column_index'] + index_number for item in items['poster']]
                        for index, km in zip(cell_index,[x['km'] for x in items['poster']]):
                            sheet[index] = km

                        break
    if 'Träd-besiktning' in items['poster'][0].keys() and any([x['Träd-besiktning'] for x in items['poster'] if x!='0']):
        for cell in excel_range_to_list("A10:A16"):
            if sheet[cell].value: 
                if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
            if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]: 
                sheet[cell] = [x["personalnamn"] for x in items["poster"] if x["personalnamn"]][0]
                index_number = cell[1:]
  
                cell_index = [item['column_index'] + index_number for item in items['poster']]
                for index, besiktartimmar in zip(cell_index,[x['Träd-besiktning'] for x in items['poster']]):
                    sheet[index] = besiktartimmar
                break
            
        if any([x['km'] for x in items['poster'] if x!='0']):
                for cell in excel_range_to_list("A18:A24"):
                    if sheet[cell].value: 
                        if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                    if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                        sheet[cell] = [x["personalnamn"] for x in items["poster"] if x["personalnamn"]][0]
                        
                        index_number = cell[1:]
                        cell_index = [item['column_index'] + index_number for item in items['poster']]
                        for index, km in zip(cell_index,[x['km'] for x in items['poster']]):
                            sheet[index] = km

                        break
                    
                    
    # SET Platschef Cells AND BYGGMÖTEN
    if 'Byggmöten' in items['poster'][0].keys() and any([x['Byggmöten'] for x in items['poster'] if x!='0']) and 'Platschef' in items['poster'][0].keys():
        for i,item in enumerate(items['poster']):
            items['poster'][i]['Platschef'] = items['poster'][i]['Platschef'] + items['poster'][i]['Byggmöten']

    if 'Platschef' in items['poster'][0].keys():

        if any([x['Platschef'] for x in items['poster'] if x!='0']):

            for cell in excel_range_to_list("A7:A8"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['Platschef'] for x in items['poster']]):
                        sheet[index] = besiktartimmar

                    break
            
        if any([x['km'] for x in items['poster'] if x!='0']):

                for cell in excel_range_to_list("A18:A24"):
                    if sheet[cell].value: 
                        if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                    if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                        sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                        index_number = cell[1:]
                        cell_index = [item['column_index'] + index_number for item in items['poster']]
                        for index, km in zip(cell_index,[x['km'] for x in items['poster']]):
                            sheet[index] = km

                        break
                    # SET MARKARBETE TIMMAR
    if 'Mark Arb' in items['poster'][0].keys():
        if any([x['Mark Arb'] for x in items['poster'] if x!='0']):

            for cell in excel_range_to_list("A69:A76"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['Mark Arb'] for x in items['poster']]):
                        sheet[index] = besiktartimmar

                    break
                        # SET SKOTARE TIMMAR
    if 'Skotare' in items['poster'][0].keys():
        if any([x['Skotare'] for x in items['poster'] if x!='0']):

            for cell in excel_range_to_list("A93:A101"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['Skotare'] for x in items['poster']]):
                        sheet[index] = besiktartimmar

                    break
                
    if 'Skotning' in items['poster'][0].keys():
        if any([x['Skotning'] for x in items['poster'] if x!='0']):

            for cell in excel_range_to_list("A93:A101"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['Skotning'] for x in items['poster']]):
                        sheet[index] = besiktartimmar

                    break                
                
    if 'Avant med förare' in items['poster'][0].keys():
        if any([x['Avant med förare'] for x in items['poster'] if x!='0']):

            for cell in excel_range_to_list("A78:A81"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['Avant med förare'] for x in items['poster']]):
                        sheet[index] = besiktartimmar

                    break
                
    if 'Lastbil' in items['poster'][0].keys():
        if any([x['Lastbil'] for x in items['poster'] if x!='0']):

            for cell in excel_range_to_list("A103:A107"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['Lastbil'] for x in items['poster']]):
                        sheet[index] = besiktartimmar
                    break
    if 'SOS-Ledare' in items['poster'][0].keys():
        if any([x['SOS-Ledare'] for x in items['poster'] if x!='0']):

            for cell in excel_range_to_list("A114:A119"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['SOS-ledare'] for x in items['poster']]):
                        sheet[index] = besiktartimmar
                    break
    if 'SOS ledare' in items['poster'][0].keys():
        if any([x['SOS ledare'] for x in items['poster'] if x!='0']):

            for cell in excel_range_to_list("A114:A119"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['SOS ledare'] for x in items['poster']]):
                        sheet[index] = besiktartimmar
                    break
                
    if 'Övrigt' in items['poster'][0].keys():
        if any([x['Övrigt'] for x in items['poster'] if x!='0']):

            for cell in excel_range_to_list("A122:A127"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['Övrigt'] for x in items['poster']]):
                        sheet[index] = besiktartimmar
                    break
    if ["Oberäknad tid" in item.keys() for item in items["poster"]]:

        if any([x['Oberäknad tid'] for x in items['poster'] if x!='0' and "Oberäknad tid" in x.keys()]):


            for cell in excel_range_to_list("A129:A145"):
                if sheet[cell].value: 
                    if "Okänd" in sheet[cell].value: unknown = sheet[cell].value+'_1'
                if not sheet[cell].value or sheet[cell].value ==[x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]:
                    sheet[cell] = [x["personalnamn"] if x["personalnamn"] else unknown for x in items["poster"]][0]
                    index_number = cell[1:]

                    cell_index = [item['column_index'] + index_number for item in items['poster']]
                    for index, besiktartimmar in zip(cell_index,[x['Oberäknad tid'] for x in items['poster']]):
                        sheet[index] = besiktartimmar
                    break
    
    if items['info']['Övrig arbetstid']:
        if sheet["AA156"].value: sheet["AA156"] = sheet["AA156"].value + items['info']['Övrig arbetstid']
        else: sheet["AA156"] = items['info']['Övrig arbetstid']
    
    
    return wb
    
    
def excel_range_to_list(range_string):

    start_cell, end_cell = range_string.split(':')
    start_col, start_row = openpyxl.utils.cell.coordinate_from_string(start_cell)
    end_col, end_row = openpyxl.utils.cell.coordinate_from_string(end_cell)
    
    cell_list = []
    for row in range(start_row, end_row+1):
        for col in range(openpyxl.utils.column_index_from_string(start_col),
                         openpyxl.utils.column_index_from_string(end_col)+1):
            cell = openpyxl.utils.cell.get_column_letter(col) + str(row)
            cell_list.append(cell)
    return cell_list

def get_date_range(start_date,end_date):
    """Creates a date range. Takes in start date and end date as inputs,
    returns a list of dates"""
    delta = datetime.timedelta(days=1)           # define the time difference between dates
    all_dates = []                      
    while start_date <= end_date:
        all_dates.append(start_date)  # add current date to the list
        start_date += delta
    return all_dates


def stringify_dict(d):
    """
    Recursively converts all keys and values of a dictionary to string data type.

    Args:
        d (dict): The dictionary to be converted.

    Returns:
        dict: The new dictionary with all keys and values converted to strings.
    """
    new_dict = {}
    for k, v in d.items():
        if isinstance(k, datetime.date):
            k = k.strftime('%Y-%m-%d')
        elif isinstance(k, datetime.time):
            k = k.strftime('%H:%M:%S')
        elif isinstance(k, (int, float)):
            k = str(k)
        if isinstance(v, dict):
            v = stringify_dict(v)
        elif isinstance(v, (list, set, tuple)):
            v = [stringify_dict(item) if isinstance(item, dict) else str(item) for item in v]
        elif isinstance(v, datetime.date):
            v = v.strftime('%Y-%m-%d')
        elif isinstance(v, datetime.time):
            v = v.strftime('%H:%M:%S')
        elif isinstance(v, (int, float)):
            v = str(v)
        new_dict[str(k)] = v
    return new_dict



if __name__ == '__main__':
    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__),'Dagböcker.xlsx'))
    for sheet in wb.worksheets:
        items = get_dictionary_from_dagbok_sheet(sheet)
        wb.close()
        wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__),'001.xlsx'))
        wb = enter_items_into_sheet(wb,items)
        items = stringify_dict(items)
        
        with open(os.path.join(os.path.join(os.path.dirname(__file__),'items'),items["info"]["Boktitel"]+' '+items["info"]["Första dag i veckan"]+'.json'),'w') as f:
            json.dump(items,f)
    
        wb.save(os.path.join(os.path.dirname(__file__),'001.xlsx'))
        wb.close()
        